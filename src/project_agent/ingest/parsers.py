from pathlib import Path

import fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook


def _is_garbled(text: str) -> bool:
    """추출된 텍스트가 깨져 있는지 확인합니다."""
    if not text:
        return True
    # 유니코드 제어문자/비정상 문자 비율 확인
    bad = sum(1 for c in text if ord(c) > 0x0FFF and ord(c) < 0x2000)
    return bad / max(len(text), 1) > 0.05


def _ocr_page_ollama(page) -> str:
    """Ollama 비전 모델을 사용해 페이지 이미지에서 텍스트를 추출합니다."""
    import base64
    import httpx
    from project_agent.config import get_settings
    settings = get_settings()
    pix = page.get_pixmap(dpi=150)
    img_b64 = base64.b64encode(pix.tobytes()).decode()
    try:
        with httpx.Client(base_url=settings.ollama_base_url, timeout=120.0, trust_env=False) as client:
            resp = client.post("/api/chat", json={
                "model": settings.ollama_model,
                "messages": [{
                    "role": "user",
                    "content": "이 이미지의 모든 텍스트를 그대로 추출하세요. 설명 없이 텍스트만 출력하세요.",
                    "images": [img_b64],
                }],
                "stream": False,
                "options": {"temperature": 0.1, "num_predict": 1024},
            })
            resp.raise_for_status()
            return resp.json()["message"]["content"].strip()
    except Exception as e:
        print(f"[OCR ERROR] Ollama vision failed: {e}")
    return ""


def parse_pdf(path: Path) -> str:
    parts: list[str] = []
    try:
        doc = fitz.open(path)
    except Exception as e:
        print(f"[PDF ERROR] Cannot open {path.name}: {e}")
        return ""
    for i, page in enumerate(doc, start=1):
        text = page.get_text().strip()
        # 텍스트가 깨져 있으면 OCR로 폴백
        if _is_garbled(text):
            ocr_text = _ocr_page_ollama(page)
            if ocr_text:
                text = ocr_text
        if text:
            parts.append(f"## Page {i}\n\n{text}")
    doc.close()
    return "\n\n".join(parts) if parts else ""


def parse_docx(path: Path) -> str:
    doc = DocxDocument(path)
    parts: list[str] = []
    for para in doc.paragraphs:
        t = para.text.strip()
        if t:
            style = para.style.name if para.style else ""
            if "Heading" in style:
                level = style.replace("Heading", "").strip() or "2"
                try:
                    n = min(int(level), 6)
                except ValueError:
                    n = 2
                parts.append(f"{'#' * n} {t}")
            else:
                parts.append(t)
    return "\n\n".join(parts)


def parse_xlsx(path: Path) -> str:
    """XLSX 파싱. 보안(암호)이 걸린 파일은 xlwings(Excel COM)로 폴백."""
    try:
        return _parse_xlsx_openpyxl(path)
    except Exception as exc:
        import logging
        logger = logging.getLogger("project_agent")
        logger.info("openpyxl failed for %s (%s), trying xlwings fallback", path.name, exc)
        try:
            return _parse_xlsx_xlwings(path)
        except Exception as exc2:
            logger.warning("xlwings also failed for %s: %s", path.name, exc2)
            raise


def _parse_xlsx_openpyxl(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}\n")
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("| " + " | ".join(cells) + " |")
        if rows and len(rows) > 1:
            header = rows[0]
            sep = "| " + " | ".join(["---"] * (rows[0].count("|") + 1)) + " |"
            parts.append("\n".join([header, sep, *rows[1:]]))
        elif rows:
            parts.append("\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def _parse_xlsx_xlwings(path: Path) -> str:
    """xlwings(Excel COM 자동화)로 보안 걸린 XLSX 파일 읽기."""
    import xlwings as xw

    app = xw.App(visible=False)
    app.display_alerts = False
    try:
        wb = app.books.open(str(path.resolve()))
        parts: list[str] = []
        for sheet in wb.sheets:
            sheet_name = sheet.name
            # 사용된 범위 전체를 읽기
            used = sheet.used_range
            if used is None:
                continue
            rows_data = used.value  # list of lists (or None)
            if not rows_data:
                continue
            parts.append(f"## Sheet: {sheet_name}\n")
            rows: list[str] = []
            for row in rows_data:
                if row is None:
                    continue
                cells = [str(c) if c is not None else "" for c in row]
                if any(c != "" for c in cells):
                    rows.append("| " + " | ".join(cells) + " |")
            if rows and len(rows) > 1:
                header = rows[0]
                sep = "| " + " | ".join(["---"] * (rows[0].count("|") + 1)) + " |"
                parts.append("\n".join([header, sep, *rows[1:]]))
            elif rows:
                parts.append("\n".join(rows))
        wb.close()
        return "\n\n".join(parts)
    finally:
        app.quit()


def parse_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def parse_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix in (".docx",):
        return parse_docx(path)
    if suffix in (".xlsx", ".xlsm"):
        return parse_xlsx(path)
    if suffix in (".txt", ".md", ".markdown", ".csv"):
        return parse_text(path)
    raise ValueError(f"Unsupported file type: {suffix}")
